import xlsxwriter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill


def create_dropdown(df, file_path, col = 5):
    color_options = ["à envoyer", "draft", "envoyé", "pas trouvé"]
    color_codes = ["#ff8e8e", "#ffeeb0", "#b2ffaf", "#daeef3"]

    workbook = xlsxwriter.Workbook(file_path, {"nan_inf_to_errors": True})
    worksheet = workbook.add_worksheet()

    # Write column headers
    worksheet.write_row("A1", df.columns)

    # Write data rows
    for idx, row in df.iterrows():
        worksheet.write_row(idx + 1, 0, row)

    
    for row in range(1, len(df) + 1):
        validation = {
            "validate": "list",
            "source": color_options,
        }

        worksheet.data_validation(row, col, row, col, validation)
        worksheet.write(row, col, color_options[0])

        for i, option in enumerate(color_options):
            color_range = f"F{row + 1}"

            worksheet.conditional_format(
                color_range,
                {
                    "type": "formula",
                    "criteria": f'=$F${row + 1}="{option}"',
                    "format": workbook.add_format({"bg_color": color_codes[i]}),
                },
            )

    workbook.close()

def is_full_name(name: str):
    words = name.split()
    for word in words:
        if word.isupper():
            return True
    return False


def verify_cell(worksheet: Worksheet):
    fill_colur = PatternFill(
        start_color="ffb162", end_color="ffb162", fill_type="solid"
    )
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == "D":
                continue
            if not cell.value or (
                cell.column_letter in ["A", "C"] and not is_full_name(cell.value)
            ):
                cell.fill = fill_colur

def save_table(df, file_path):
    create_dropdown(df, file_path)
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    verify_cell(worksheet)

    table_style = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table = Table(
        displayName="MyTable", ref=worksheet.dimensions, tableStyleInfo=table_style
    )
    worksheet.add_table(table)
    for col_num, _ in enumerate(df.columns, start=0):
        worksheet.column_dimensions[
            xlsxwriter.utility.xl_col_to_name(col_num)
        ].width = 30

    workbook.save(file_path)
